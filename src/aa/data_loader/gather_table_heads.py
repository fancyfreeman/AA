import pandas as pd
from openpyxl import load_workbook

def gather_table_heads():
    # 读取配置文件
    config_file = "/Users/chenxin/Dev/AA/config/data_extraction_config.xlsx"
    config_df = pd.read_excel(config_file, sheet_name='multi_sheet_df')

    # 初始化输出文件
    output_file = "/Users/chenxin/Dev/AA/data/processed/table_heads.xlsx"
    writer = pd.ExcelWriter(output_file, engine='openpyxl')

    # 遍历每个sheet_name
    for sheet_name in config_df['sheet_name'].unique():
        sheet_df = config_df[config_df['sheet_name'] == sheet_name]

        # 遍历每个file_name
        for _, row in sheet_df.iterrows():
            file_name = f"{row['file_name']}.xlsx"
            start_row = row['start_row']
            end_row = row['end_row']

            # 读取原始文件
            input_file = f"/Users/chenxin/Dev/AA/data/raw/{file_name}"
            wb = load_workbook(input_file)
            ws = wb[sheet_name]

            # 获取表头区域
            header_rows = []
            for i in range(1, start_row + 1):
                header_rows.append([cell.value for cell in ws[i]])
            for i in range(end_row-2, end_row + 1):
                header_rows.append([cell.value for cell in ws[i]])

            # 将表头写入输出文件
            if sheet_name in writer.book.sheetnames:
                output_ws = writer.book[sheet_name]
                start_row = output_ws.max_row + 2  # 从新行开始，留空一行
            else:
                output_ws = writer.book.create_sheet(sheet_name)
                start_row = 1

            for row_idx, row_data in enumerate(header_rows):
                for col_idx, value in enumerate(row_data):
                    output_ws.cell(row=start_row+row_idx, column=col_idx+1, value=value)

            # 添加空行
            output_ws.append([])

    # 保存输出文件
    # writer.save(output_file)
    writer.close()
if __name__ == '__main__':
    gather_table_heads()
